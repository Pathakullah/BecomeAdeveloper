import os

from openpyxl import *
from os import path

#wb_path
#naal


wb = load_workbook("Fluke Report.xlsx")
ws1 = wb["Data"]
ws2 = wb["Summary"]
ws3 = wb["BK"]
max_data = ws1.max_row

i=2
j = ws2.max_row+1



while i <= max_data:
    if ws1.cell(i,521).value != "Converted" :
        Cable_Id = ws1.cell(i,1).value
        naal = ws1.cell(i,14).value
        cable_type = ws1.cell(i,12).value
        test_std = ws1.cell(i,10).value
        Len_mtr = ws1.cell(i,18).value
        Over_Result = ws1.cell(i,13).value

        # return loss value get
        R1_N = ws1.cell(i, 229).value
        R2_N = ws1.cell(i, 232).value
        R3_N = ws1.cell(i, 235).value
        R4_N = ws1.cell(i, 238).value
        R1_F = ws1.cell(i, 241).value
        R2_F = ws1.cell(i, 244).value
        R3_F = ws1.cell(i, 247).value
        R4_F = ws1.cell(i, 250).value
        # NEXT VALUE GET IT
        N1_N = ws1.cell(i, 109).value
        N2_N = ws1.cell(i, 112).value
        N3_N = ws1.cell(i, 115).value
        N4_N = ws1.cell(i, 118).value
        N5_N = ws1.cell(i, 121).value
        N6_N = ws1.cell(i, 124).value
        N1_F = ws1.cell(i, 127).value
        N2_F = ws1.cell(i, 130).value
        N3_F = ws1.cell(i, 133).value
        N4_F = ws1.cell(i, 136).value
        N5_F = ws1.cell(i, 139).value
        N6_F = ws1.cell(i, 142).value

        IL1 = ws1.cell(i, 97).value
        IL2 = ws1.cell(i, 100).value
        IL3 = ws1.cell(i, 103).value
        IL4 = ws1.cell(i, 106).value

        RL = ws1.cell(i, 63).value
        IL = ws1.cell(i, 53).value
        NXT = ws1.cell(i,55).value
        Result = ws1.cell(i, 13).value
        Resistance = ws1.cell(i, 45).value

        min_r1 = (min(R1_N, R1_F))
        min_r2 = (min(R2_N, R2_F))
        min_r3 = (min(R3_N, R3_F))
        min_r4 = (min(R4_N, R4_F))
        # 29,23
        min_n1 = (min(N1_N, N1_F))
        min_n2 = (min(N2_N, N2_F))
        min_n3 = (min(N3_N, N3_F))
        min_n4 = (min(N4_N, N4_F))
        min_n5 = (min(N5_N, N5_F))
        min_n6 = (min(N6_N, N6_F))

        #------------------------------------------------------------
        ws2.cell(j, 1).value = Cable_Id
        ws2.cell(j, 2).value = naal
        ws2.cell(j, 3).value = cable_type
        ws2.cell(j, 4).value = test_std
        ws2.cell(j, 5).value = Len_mtr
        ws2.cell(j, 6).value = Over_Result

        ws2.cell(j, 37).value = R1_N
        ws2.cell(j, 38).value = R2_N
        ws2.cell(j, 39).value = R3_N
        ws2.cell(j, 40).value = R4_N
        ws2.cell(j, 41).value = R1_F
        ws2.cell(j, 42).value = R2_F
        ws2.cell(j, 43).value = R3_F
        ws2.cell(j, 44).value = R4_F

        ws2.cell(j, 45).value = N1_N
        ws2.cell(j, 46).value = N2_N
        ws2.cell(j, 47).value = N3_N
        ws2.cell(j, 48).value = N4_N
        ws2.cell(j, 49).value = N5_N
        ws2.cell(j, 50).value = N6_N

        ws2.cell(j, 51).value = N1_F
        ws2.cell(j, 52).value = N2_F
        ws2.cell(j, 53).value = N3_F
        ws2.cell(j, 54).value = N4_F
        ws2.cell(j, 55).value = N5_F
        ws2.cell(j, 56).value = N6_F

        ws2.cell(j, 33).value = IL1
        ws2.cell(j, 34).value = IL2
        ws2.cell(j, 35).value = IL3
        ws2.cell(j, 36).value = IL4

        ws2.cell(j, 20).value = RL
        ws2.cell(j, 21).value = IL
        ws2.cell(j, 22).value = NXT
        ws2.cell(j, 19).value = Result
        ws2.cell(j, 57).value = Resistance

        ws2.cell(j, 29).value = min_r1
        ws2.cell(j, 30).value = min_r2
        ws2.cell(j, 31).value = min_r3
        ws2.cell(j, 32).value = min_r4

        ws2.cell(j, 23).value = min_n1
        ws2.cell(j, 24).value = min_n2
        ws2.cell(j, 25).value = min_n3
        ws2.cell(j, 26).value = min_n4
        ws2.cell(j, 27).value = min_n5
        ws2.cell(j, 28).value = min_n6


        ws1.cell(i, 521).value = "Converted"
        j = j + 1
    i = i+1


print("File has been converted")
wb.save("Fluke Report.xlsx")
